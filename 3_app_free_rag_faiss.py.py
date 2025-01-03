import streamlit as st
import requests, re, json
import os
from typing import Any
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from langchain_community.vectorstores import FAISS
from langchain_community.embeddings import HuggingFaceEmbeddings, OpenAIEmbeddings
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.schema import Document

# システムコンテンツの定義
SYSTEM_CONTENT = """
You are a helpful, respectful and honest assistant. Always answer as helpfully as possible, while being safe.  Your answers should not include any harmful, unethical, racist, sexist, toxic, dangerous, or illegal content. Please ensure that your responses are socially unbiased and positive in nature.

If a question does not make any sense, or is not factually coherent, explain why instead of answering something not correct. If you don't know the answer to a question, please don't share false information.

You must answer in Japanese.
"""

chat_url = 'https://bam-api.res.ibm.com/v2/text/chat_stream?version=2024-05-10'
chat_model = 'meta-llama/llama-3-70b-instruct'
# embed_url = 'https://bam-api.res.ibm.com/v2/text/embeddings?version=2024-04-15'
embed_model = 'intfloat/multilingual-e5-large'
headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer pak-JcUU6jMIwSl9xJtO-bV-IvXbuU3sUZufLfMvWo5ThXc'
}

# フォルダが存在しない場合は作成
db_folder = 'db'
static_folder = 'static'
def create_folders():
    folders = [db_folder, static_folder]
    for folder in folders:
        if not os.path.exists(folder):
            os.makedirs(folder)

# 日本語のテキストスプリッターの定義
class JapaneseCharacterTextSplitter(RecursiveCharacterTextSplitter):
    def __init__(self, **kwargs: Any):
        separators = ["\n\n", "\n", "。"]
        super().__init__(separators=separators, **kwargs)

# ファイルアップロード
def store_files(uploaded_files):
    for uploaded_file in uploaded_files:
        file_path = os.path.join(static_folder, uploaded_file.name)
        with open(file_path, 'wb') as f:
            f.write(uploaded_file.getbuffer())
        st.info(f"{file_path} に保存されました。")

# ファイルの読み取りとテキスト抽出
def read_file(file_path):
    _, file_extension = os.path.splitext(file_path)
    text = ""
    if file_extension == '.txt':
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()
    elif file_extension == '.pdf':
        reader = PdfReader(file_path)
        for page in reader.pages:
            text += page.extract_text()
    elif file_extension == '.xlsx':
        workbook = load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text = ' '.join([str(cell) for cell in row if cell is not None])
                text += row_text + '\n'
    if text:
        st.info(f"{file_path} を読み込みました")
    return text

# ベクトルストアの作成
def create_vector_store():
    all_texts = []
    for root, _, files in os.walk(static_folder):
        for file_name in files:
            file_path = os.path.join(root, file_name)
            text = read_file(file_path)
            all_texts.append(text)

    if not all_texts:
        st.error("有効なテキストが見つかりませんでした。")
        return

    # テキスト分割
    splitter = JapaneseCharacterTextSplitter(chunk_size=2000, chunk_overlap=400)
    documents = [Document(page_content=chunk) for text in all_texts for chunk in splitter.split_text(text)]

    # ベクトルストアの作成
    embeddings = HuggingFaceEmbeddings(model_name=embed_model)
    vector_store = FAISS.from_documents(documents, embeddings)
    vector_store.save_local(db_folder)
    st.success("ベクトルストアの作成が完了しました。")

# Chroma DBの中身を一覧表示
def list_chroma_db_contents():
    embeddings = HuggingFaceEmbeddings(model_name=embed_model)
    vector_store = FAISS.load_local(db_folder, embeddings, allow_dangerous_deserialization=True)
    
    # 全てのドキュメントを取得
    documents = vector_store.docstore._dict
    
    # ドキュメントを表示
    st.write(f"Number of chunks: {len(documents)}")
    st.write(documents)

# 質問の入力
def input_question(question):
    st.session_state.question = st.text_input("入力", question)

# 質問への回答
def answer(input):
    # Stream
    data = {
        "model_id": chat_model,
        "messages": [
            {"role": "system", "content": SYSTEM_CONTENT},
            {"role": "user", "content": input}
        ],
        "parameters": {
            "decoding_method": "greedy",
            "min_new_tokens": 1,
            "max_new_tokens": 1000
        }
    }

    print(f"Query: {SYSTEM_CONTENT}")
    print(f"{input}")
    # ストリーミング応答の処理
    response_text = ''
    data_pattern = re.compile(rb'^data:\s*(.*)')
    response = requests.post(chat_url, headers=headers, json=data, stream=True)
    for line in response.iter_lines():
        match = data_pattern.match(line)
        if match:
            json_data = match.group(1).decode('utf-8')
            json_loaded = json.loads(json_data)
            content = json_loaded["results"][0]["generated_text"]
            response_text += content
            yield content
    print(f"Output: {response_text}")

# RAGを参照した質問への回答
def answer_with_rag(input):
    # Chromaベクトルストアをロード
    embeddings = HuggingFaceEmbeddings(model_name=embed_model)
    vector_store = FAISS.load_local(db_folder, embeddings, allow_dangerous_deserialization=True)

    # 類似度検索
    results = vector_store.similarity_search_with_score(input, k=2)

    # 検索結果と類似度を表示
    for doc, score in results:
        # print(f"Document: {doc.page_content[:200]}...")  # ドキュメントの最初の200文字を表示
        print(f"Document: {doc}")  # ドキュメントの最初の200文字を表示
        print(f"Score: {score}")

    # 検索結果を結合してコンテキストとして使用
    context = "\n".join([doc.page_content for doc, _ in results])

    query = f"""
回答にあたり、以下の文脈を考慮してください。
<文脈>
{context}

ユーザーの質問は以下です:
{input}

なお、必ず日本語で回答してください。
"""
    return answer(query)

# Streamlitのページ設定とタイトル
st.set_page_config(layout="wide")
st.title("RAG Chatbot Sample")

# フォルダ作成
create_folders()

# ファイルアップロード
uploaded_files = st.file_uploader("1. Vector Storeを作成するファイルを全てアップロードしてください。", accept_multiple_files=True)
if uploaded_files:
    with st.spinner("実行中…"):
        store_files(uploaded_files)

# ベクトルストア作成
if st.button("2. ベクトルストア作成"):
    with st.spinner("実行中…"):
        create_vector_store()

# ベクトルストアの中身を一覧表示
if st.button("3. ベクトルストアの中身を一覧表示"):
    with st.spinner("実行中…"):
        list_chroma_db_contents()

st.write("---")
st.write("こちらはチャットボットのサンプルです。1問1答形式で、会話の履歴は覚えていません。")

# 質問の入力
input_question("カレーのトッピングでおすすめは?")

# フリー生成ボタン
if st.button('フリー生成'):
    st.write("---")
    with st.spinner("実行中…"):
        st.write_stream(answer(st.session_state.question))

# RAG参照生成ボタン
if st.button('RAG参照生成'):
    st.write("---")
    with st.spinner("実行中…"):
        # Stream
        st.write_stream(answer_with_rag(st.session_state.question))
